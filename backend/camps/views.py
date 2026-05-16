import re

from django.db import transaction
from openpyxl import load_workbook
from rest_framework import serializers, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.viewsets import ModelViewSet

from competitions.models import IntegralRound

from .models import Camp, CampEnrollment, Coach, Judge, Student, Team
from .permissions import IsAdminOrStaff
from .serializers import (
    CampEnrollmentSerializer,
    CampSerializer,
    CoachSerializer,
    JudgeSerializer,
    StudentSerializer,
    TeamSerializer,
)


class StaffModelViewSet(ModelViewSet):
    permission_classes = [IsAdminOrStaff]


class CampViewSet(StaffModelViewSet):
    queryset = Camp.objects.all()
    serializer_class = CampSerializer

    def perform_create(self, serializer):
        camp = serializer.save()
        for number in [1, 2, 3]:
            IntegralRound.objects.get_or_create(camp=camp, number=number)

    @action(detail=True, methods=["post"])
    def import_enrollments(self, request, pk=None):
        camp = self.get_object()
        uploaded_file = request.FILES.get("file")
        should_commit = str(request.data.get("commit", "false")).lower() == "true"
        if not uploaded_file:
            raise serializers.ValidationError("请上传 Excel 文件。")

        rows = _parse_enrollment_workbook(uploaded_file)
        preview = _validate_enrollment_rows(camp, rows)
        if preview["errors"]:
            return Response(preview, status=status.HTTP_400_BAD_REQUEST)

        if should_commit:
            with transaction.atomic():
                for row in preview["rows"]:
                    student = None
                    if row["matched_student_id"]:
                        student = Student.objects.get(id=row["matched_student_id"])
                    else:
                        student = Student.objects.create(real_name=row["real_name"], phone=row["phone"])
                    CampEnrollment.objects.create(
                        camp=camp,
                        student=student,
                        nickname=row["nickname"],
                    )
            preview["committed"] = True
        else:
            preview["committed"] = False

        return Response(preview)


class CoachViewSet(StaffModelViewSet):
    queryset = Coach.objects.all()
    serializer_class = CoachSerializer


class JudgeViewSet(StaffModelViewSet):
    queryset = Judge.objects.all()
    serializer_class = JudgeSerializer


class StudentViewSet(StaffModelViewSet):
    queryset = Student.objects.all()
    serializer_class = StudentSerializer


class TeamViewSet(StaffModelViewSet):
    queryset = Team.objects.select_related("camp", "coach").prefetch_related("members")
    serializer_class = TeamSerializer

    @action(detail=False, methods=["post"])
    def create_from_nicknames(self, request):
        camp_id = request.data.get("camp")
        coach_id = request.data.get("coach")
        team_name = str(request.data.get("name", "")).strip()
        raw_nicknames = str(request.data.get("nicknames", ""))

        if not camp_id or not coach_id or not team_name:
            raise serializers.ValidationError("请填写营期、队名和教练。")

        nicknames = [line.strip() for line in raw_nicknames.replace("，", "\n").replace(",", "\n").splitlines() if line.strip()]
        errors = []
        if len(nicknames) not in {7, 8}:
            errors.append("队伍必须一次录入 7-8 个学员昵称。")
        if len(nicknames) != len(set(nicknames)):
            errors.append("输入的昵称中有重复。")
        if Team.objects.filter(camp_id=camp_id, name=team_name).exists():
            errors.append("本期已存在同名队伍。")
        if Team.objects.filter(camp_id=camp_id, coach_id=coach_id).exists():
            errors.append("该教练本期已经带队。")

        enrollments = list(CampEnrollment.objects.filter(camp_id=camp_id, nickname__in=nicknames).select_related("student", "team"))
        enrollment_by_nickname = {enrollment.nickname: enrollment for enrollment in enrollments}
        matched = []
        for nickname in nicknames:
            enrollment = enrollment_by_nickname.get(nickname)
            if not enrollment:
                errors.append(f"未找到本期学员昵称：{nickname}")
                continue
            if enrollment.team_id:
                errors.append(f"{nickname} 已在队伍 {enrollment.team.name} 中。")
                continue
            matched.append(
                {
                    "id": enrollment.id,
                    "nickname": enrollment.nickname,
                    "student_name": enrollment.student.real_name,
                }
            )

        if errors:
            return Response({"errors": errors, "matched": matched}, status=status.HTTP_400_BAD_REQUEST)

        with transaction.atomic():
            team = Team.objects.create(camp_id=camp_id, coach_id=coach_id, name=team_name)
            CampEnrollment.objects.filter(id__in=[item["id"] for item in matched]).update(team=team)

        return Response(
            {
                "team": TeamSerializer(team).data,
                "matched": matched,
            },
            status=status.HTTP_201_CREATED,
        )


class CampEnrollmentViewSet(StaffModelViewSet):
    queryset = CampEnrollment.objects.select_related("camp", "student", "team")
    serializer_class = CampEnrollmentSerializer

# Create your views here.


def _normalize_phone(value):
    if value is None:
        return ""
    text = str(value).strip()
    return re.sub(r"\D", "", text)


def _parse_enrollment_workbook(uploaded_file):
    workbook = load_workbook(uploaded_file, read_only=True, data_only=True)
    worksheet = workbook.worksheets[0]
    rows = []
    for index, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
        if index == 1:
            continue
        real_name = str(row[0]).strip() if len(row) > 0 and row[0] is not None else ""
        nickname = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
        phone = _normalize_phone(row[2] if len(row) > 2 else "")
        if not real_name and not nickname and not phone:
            continue
        rows.append({"row": index, "real_name": real_name, "nickname": nickname, "phone": phone})
    return rows


def _validate_enrollment_rows(camp, rows):
    errors = []
    seen_nicknames = {}
    seen_phones = {}
    existing_nicknames = set(CampEnrollment.objects.filter(camp=camp).values_list("nickname", flat=True))
    existing_by_phone = {student.phone: student for student in Student.objects.exclude(phone="") if student.phone}
    existing_enrollment_student_ids = set(CampEnrollment.objects.filter(camp=camp).values_list("student_id", flat=True))
    preview_rows = []
    new_student_count = 0
    matched_student_count = 0

    for row in rows:
        row_errors = []
        if not row["real_name"]:
            row_errors.append("真实姓名不能为空。")
        if not row["nickname"]:
            row_errors.append("本期昵称不能为空。")
        if not row["phone"]:
            row_errors.append("电话不能为空。")

        if row["nickname"]:
            previous_row = seen_nicknames.get(row["nickname"])
            if previous_row:
                row_errors.append(f"本次 Excel 内昵称重复，首次出现在第 {previous_row} 行。")
            if row["nickname"] in existing_nicknames:
                row_errors.append("本期已存在相同昵称。")
            seen_nicknames[row["nickname"]] = row["row"]

        if row["phone"]:
            previous = seen_phones.get(row["phone"])
            if previous and previous["real_name"] != row["real_name"]:
                row_errors.append(f"本次 Excel 内同一电话对应不同真实姓名，首次出现在第 {previous['row']} 行。")
            seen_phones[row["phone"]] = row

        matched_student = existing_by_phone.get(row["phone"])
        matched_student_id = None
        matched_status = "new"
        if matched_student:
            matched_student_id = matched_student.id
            if matched_student.real_name != row["real_name"]:
                row_errors.append("电话已存在于总学员库，但真实姓名不同。")
            elif matched_student.id in existing_enrollment_student_ids:
                row_errors.append("该学员已在本期报名。")
            else:
                matched_status = "matched"

        if row_errors:
            errors.append({"row": row["row"], "messages": row_errors})
        else:
            if matched_status == "matched":
                matched_student_count += 1
            else:
                new_student_count += 1

        preview_rows.append(
            {
                **row,
                "matched_student_id": matched_student_id,
                "status": matched_status,
                "errors": row_errors,
            }
        )

    return {
        "total": len(rows),
        "new_student_count": new_student_count,
        "matched_student_count": matched_student_count,
        "rows": preview_rows,
        "errors": errors,
    }
