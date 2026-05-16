from django.http import HttpResponse
from openpyxl import Workbook
from rest_framework.response import Response
from rest_framework.views import APIView

from accounts.models import Role
from accounts.serializers import UserAccountSerializer
from assessments.models import AssessmentAssignment, AssessmentVenue
from assessments.serializers import AssessmentAssignmentSerializer, AssessmentVenueSerializer
from camps.models import Camp, CampEnrollment, Coach, Judge, Student, Team
from camps.permissions import IsAdminOrStaff
from camps.serializers import CampEnrollmentSerializer, CampSerializer, CoachSerializer, JudgeSerializer, StudentSerializer, TeamSerializer
from competitions.models import CompetitionVenue, DebateSide, IntegralRound, Match
from competitions.serializers import CompetitionVenueSerializer, IntegralRoundSerializer, MatchSerializer


def _match_review_summary(match):
    assigned_judges = list(match.venue.judges.all())
    ballots = list(match.ballots.all())
    submitted_ballots = [ballot for ballot in ballots if ballot.submitted_at]
    draft_ballots = [ballot for ballot in ballots if not ballot.submitted_at]
    positions = list(match.positions.all())
    positions_by_id = {position.id: position for position in positions}
    affirmative_position_count = sum(1 for position in positions if position.side == DebateSide.AFFIRMATIVE)
    negative_position_count = sum(1 for position in positions if position.side == DebateSide.NEGATIVE)
    expected_score_count = len(assigned_judges) * len(positions)
    submitted_score_count = sum(ballot.position_scores.count() for ballot in submitted_ballots)
    position_score_totals = {DebateSide.AFFIRMATIVE: 0.0, DebateSide.NEGATIVE: 0.0}
    best_speaker_totals = {}
    ballot_details = []
    position_details = [
        {
            "id": position.id,
            "side": position.side,
            "position_number": position.position_number,
            "label": position.display_name,
            "speaker": position.enrollment.nickname,
            "student_name": position.enrollment.student.real_name,
            "team": position.enrollment.team_id,
            "coach_note": position.coach_note,
        }
        for position in positions
    ]

    for ballot in submitted_ballots:
        scores = []
        for score in ballot.position_scores.all():
            position = positions_by_id.get(score.position_id)
            if position:
                position_score_totals[position.side] += float(score.score)
                scores.append(
                    {
                        "position": score.position_id,
                        "label": position.display_name,
                        "speaker": position.enrollment.nickname,
                        "side": position.side,
                        "score": float(score.score),
                        "speech_record": score.speech_record,
                        "judge_feedback": score.judge_feedback,
                    }
                )
        best_votes = []
        for vote in ballot.best_speaker_votes.all():
            position = positions_by_id.get(vote.position_id)
            if not position:
                continue
            best_votes.append(
                {
                    "position": vote.position_id,
                    "label": position.display_name,
                    "speaker": position.enrollment.nickname,
                    "side": position.side,
                    "weight": vote.weight,
                }
            )
            current = best_speaker_totals.setdefault(
                vote.position_id,
                {
                    "position": vote.position_id,
                    "label": position.display_name,
                    "speaker": position.enrollment.nickname,
                    "side": position.side,
                    "votes": 0,
                },
            )
            current["votes"] += vote.weight
        ballot_details.append(
            {
                "id": ballot.id,
                "judge": ballot.judge_id,
                "judge_name": ballot.judge.name,
                "submitted_at": ballot.submitted_at.isoformat() if ballot.submitted_at else None,
                "affirmative_votes": ballot.affirmative_votes,
                "negative_votes": ballot.negative_votes,
                "corrected_by_staff": ballot.corrected_by_staff,
                "correction_note": ballot.correction_note,
                "position_scores": sorted(scores, key=lambda item: (item["side"], item["label"])),
                "best_speaker_votes": sorted(best_votes, key=lambda item: item["weight"], reverse=True),
            }
        )

    affirmative_votes = sum(ballot.affirmative_votes for ballot in submitted_ballots)
    negative_votes = sum(ballot.negative_votes for ballot in submitted_ballots)
    affirmative_position_score = position_score_totals[DebateSide.AFFIRMATIVE]
    negative_position_score = position_score_totals[DebateSide.NEGATIVE]
    affirmative_points = round(affirmative_position_score * 0.15 + affirmative_votes, 2)
    negative_points = round(negative_position_score * 0.15 + negative_votes, 2)

    return {
        "match": match.id,
        "round_number": match.integral_round.number,
        "venue_name": match.venue.name,
        "sequence": match.sequence,
        "starts_at": match.starts_at.strftime("%H:%M"),
        "topic": match.integral_round.topic,
        "affirmative_team": match.affirmative_team_id,
        "affirmative_team_name": match.affirmative_team.name,
        "negative_team": match.negative_team_id,
        "negative_team_name": match.negative_team.name,
        "is_verified": match.is_verified,
        "verified_at": match.verified_at.isoformat() if match.verified_at else None,
        "verification_note": match.verification_note,
        "best_speaker_override": match.best_speaker_override_id,
        "assigned_judge_count": len(assigned_judges),
        "assigned_judge_names": [judge.name for judge in assigned_judges],
        "affirmative_position_count": affirmative_position_count,
        "negative_position_count": negative_position_count,
        "positions_are_complete": affirmative_position_count == 4 and negative_position_count == 4,
        "submitted_ballot_count": len(submitted_ballots),
        "draft_ballot_count": len(draft_ballots),
        "pending_ballot_count": max(len(assigned_judges) - len(ballots), 0),
        "expected_score_count": expected_score_count,
        "submitted_score_count": submitted_score_count,
        "is_complete": len(assigned_judges) > 0
        and len(submitted_ballots) == len(assigned_judges)
        and submitted_score_count == expected_score_count,
        "affirmative_votes": affirmative_votes,
        "negative_votes": negative_votes,
        "affirmative_position_score": round(affirmative_position_score, 1),
        "negative_position_score": round(negative_position_score, 1),
        "affirmative_points": affirmative_points,
        "negative_points": negative_points,
        "best_speaker_totals": sorted(best_speaker_totals.values(), key=lambda item: item["votes"], reverse=True),
        "positions": position_details,
        "ballots": ballot_details,
    }


def _student_histories():
    histories = {}
    enrollments = CampEnrollment.objects.select_related("camp", "student", "team").order_by("-camp__starts_on", "-camp_id")
    for enrollment in enrollments:
        current = histories.setdefault(
            enrollment.student_id,
            {
                "student": enrollment.student_id,
                "real_name": enrollment.student.real_name,
                "phone": enrollment.student.phone,
                "participations": [],
            },
        )
        current["participations"].append(
            {
                "camp": enrollment.camp_id,
                "camp_name": enrollment.camp.name,
                "nickname": enrollment.nickname,
                "team_name": enrollment.team.name if enrollment.team else None,
            }
        )
    return sorted(histories.values(), key=lambda item: item["real_name"])


def _student_match_stats(match_reviews):
    stats = {}
    for review in match_reviews:
        best_votes_by_position = {item["position"]: item["votes"] for item in review["best_speaker_totals"]}
        score_totals = {}
        score_counts = {}
        for ballot in review["ballots"]:
            for score in ballot["position_scores"]:
                score_totals[score["position"]] = score_totals.get(score["position"], 0) + score["score"]
                score_counts[score["position"]] = score_counts.get(score["position"], 0) + 1
        for position in review["positions"]:
            current = stats.setdefault(
                position["id"],
                {
                    "position": position["id"],
                    "speaker": position["speaker"],
                    "student_name": position["student_name"],
                    "team": position["team"],
                    "team_name": review["affirmative_team_name"] if position["side"] == DebateSide.AFFIRMATIVE else review["negative_team_name"],
                    "round_number": review["round_number"],
                    "match": review["match"],
                    "side": position["side"],
                    "label": position["label"],
                    "average_score": 0,
                    "best_speaker_votes": best_votes_by_position.get(position["id"], 0),
                },
            )
            count = score_counts.get(position["id"], 0)
            if count:
                current["average_score"] = round(score_totals[position["id"]] / count, 1)
    return sorted(stats.values(), key=lambda item: (item["round_number"], item["speaker"]))


def _active_camp():
    camp = Camp.objects.filter(is_active=True).order_by("-starts_on", "-id").first()
    if not camp:
        camp = Camp.objects.order_by("-starts_on", "-id").first()
    return camp


def _camp_data(camp):
    rounds = IntegralRound.objects.none()
    venues = CompetitionVenue.objects.none()
    matches = Match.objects.none()
    teams = Team.objects.none()
    enrollments = CampEnrollment.objects.none()

    if camp:
        rounds = IntegralRound.objects.filter(camp=camp).prefetch_related("matches")
        venues = CompetitionVenue.objects.filter(integral_round__camp=camp).prefetch_related("judges")
        matches = Match.objects.filter(integral_round__camp=camp).select_related(
            "integral_round",
            "venue",
            "affirmative_team",
            "negative_team",
            "best_speaker_override",
        ).prefetch_related(
            "venue__judges",
            "positions",
            "positions__enrollment",
            "positions__enrollment__student",
            "ballots",
            "ballots__judge",
            "ballots__position_scores",
            "ballots__best_speaker_votes",
        )
        teams = Team.objects.filter(camp=camp).select_related("coach").prefetch_related("members")
        enrollments = CampEnrollment.objects.filter(camp=camp).select_related("student", "team")
    return rounds, venues, matches, teams, enrollments


def _team_rankings(teams, match_reviews):
    team_totals = {}
    for team in teams:
        team_totals[team.id] = {
            "team": team.id,
            "team_name": team.name,
            "round_scores": {
                number: {"position_score": 0, "votes": 0, "score": 0}
                for number in [1, 2, 3]
            },
            "total": 0,
        }
    for review in match_reviews:
        if review["affirmative_team"] in team_totals:
            score = team_totals[review["affirmative_team"]]["round_scores"][review["round_number"]]
            score["position_score"] += review["affirmative_position_score"]
            score["votes"] += review["affirmative_votes"]
            score["score"] += review["affirmative_points"]
        if review["negative_team"] in team_totals:
            score = team_totals[review["negative_team"]]["round_scores"][review["round_number"]]
            score["position_score"] += review["negative_position_score"]
            score["votes"] += review["negative_votes"]
            score["score"] += review["negative_points"]
    for total in team_totals.values():
        total["round_scores"] = [
            {
                "round_number": number,
                "position_score": round(total["round_scores"][number]["position_score"], 1),
                "votes": total["round_scores"][number]["votes"],
                "score": round(total["round_scores"][number]["score"], 2),
            }
            for number in [1, 2, 3]
        ]
        total["total"] = round(sum(item["score"] for item in total["round_scores"]), 2)
    return sorted(team_totals.values(), key=lambda item: item["total"], reverse=True)


def _workbook_response(workbook, filename):
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    workbook.save(response)
    return response


class OperationsDashboardView(APIView):
    permission_classes = [IsAdminOrStaff]

    def get(self, request):
        camp = _active_camp()
        rounds, venues, matches, teams, enrollments = _camp_data(camp)
        assessment_venues = AssessmentVenue.objects.none()
        assessment_assignments = AssessmentAssignment.objects.none()
        if camp:
            assessment_venues = AssessmentVenue.objects.filter(camp=camp).prefetch_related("coaches")
            assessment_assignments = AssessmentAssignment.objects.filter(venue__camp=camp).select_related(
                "venue",
                "enrollment",
                "enrollment__student",
            )
        match_reviews = [_match_review_summary(match) for match in matches]
        team_rankings = _team_rankings(teams, match_reviews)

        users = request.user.__class__.objects.none()
        if getattr(request.user.profile, "role", None) == Role.ADMIN:
            users = request.user.__class__.objects.select_related("profile", "profile__coach", "profile__judge").order_by("username")

        return Response(
            {
                "activeCamp": CampSerializer(camp).data if camp else None,
                "camps": CampSerializer(Camp.objects.all(), many=True).data,
                "rounds": IntegralRoundSerializer(rounds, many=True).data,
                "venues": CompetitionVenueSerializer(venues, many=True).data,
                "matches": MatchSerializer(matches, many=True).data,
                "teams": TeamSerializer(teams, many=True).data,
                "coaches": CoachSerializer(Coach.objects.filter(is_active=True), many=True).data,
                "judges": JudgeSerializer(Judge.objects.filter(is_active=True), many=True).data,
                "students": StudentSerializer(Student.objects.all(), many=True).data,
                "enrollments": CampEnrollmentSerializer(enrollments, many=True).data,
                "users": UserAccountSerializer(users, many=True).data,
                "matchReviews": match_reviews,
                "teamRankings": team_rankings,
                "studentHistories": _student_histories(),
                "studentMatchStats": _student_match_stats(match_reviews),
                "assessmentVenues": AssessmentVenueSerializer(assessment_venues, many=True).data,
                "assessmentAssignments": AssessmentAssignmentSerializer(assessment_assignments, many=True).data,
            }
        )


class TeamRankingExportView(APIView):
    permission_classes = [IsAdminOrStaff]

    def get(self, request):
        camp = _active_camp()
        _, _, matches, teams, _ = _camp_data(camp)
        reviews = [_match_review_summary(match) for match in matches]
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "队伍积分榜"
        sheet.append(["排名", "队伍", "积分赛1辩位分", "积分赛1投票", "积分赛1总分", "积分赛2辩位分", "积分赛2投票", "积分赛2总分", "积分赛3辩位分", "积分赛3投票", "积分赛3总分", "总分"])
        for index, team in enumerate(_team_rankings(teams, reviews), start=1):
            row = [index, team["team_name"]]
            for round_score in team["round_scores"]:
                row.extend([round_score["position_score"], round_score["votes"], round_score["score"]])
            row.append(team["total"])
            sheet.append(row)
        return _workbook_response(workbook, "team-rankings.xlsx")


class StudentStatsExportView(APIView):
    permission_classes = [IsAdminOrStaff]

    def get(self, request):
        camp = _active_camp()
        _, _, matches, _, _ = _camp_data(camp)
        reviews = [_match_review_summary(match) for match in matches]
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "学员赛事数据"
        sheet.append(["学员昵称", "真实姓名", "所在队伍", "轮次", "辩位", "个人平均分", "最佳辩手票"])
        team_names = {position["id"]: "" for review in reviews for position in review["positions"]}
        for review in reviews:
            for position in review["positions"]:
                team_names[position["id"]] = review["affirmative_team_name"] if position["side"] == DebateSide.AFFIRMATIVE else review["negative_team_name"]
        for item in _student_match_stats(reviews):
            sheet.append([
                item["speaker"],
                item["student_name"],
                team_names.get(item["position"], ""),
                f"积分赛{item['round_number']}",
                item["label"],
                item["average_score"],
                item["best_speaker_votes"],
            ])
        return _workbook_response(workbook, "student-match-stats.xlsx")


class JudgeRecordsExportView(APIView):
    permission_classes = [IsAdminOrStaff]

    def get(self, request):
        camp = _active_camp()
        _, _, matches, _, _ = _camp_data(camp)
        reviews = [_match_review_summary(match) for match in matches]
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "评委记录汇总"
        sheet.append(["轮次", "会场", "场次", "评委", "辩位", "学员", "分数", "发言记录", "评委反馈", "最终正方票", "最终反方票", "最佳辩手票"])
        for review in reviews:
            for ballot in review["ballots"]:
                best_votes = {vote["position"]: vote["weight"] for vote in ballot["best_speaker_votes"]}
                for score in ballot["position_scores"]:
                    sheet.append([
                        f"积分赛{review['round_number']}",
                        review["venue_name"],
                        review["sequence"],
                        ballot["judge_name"],
                        score["label"],
                        score["speaker"],
                        score["score"],
                        score["speech_record"],
                        score["judge_feedback"],
                        ballot["affirmative_votes"],
                        ballot["negative_votes"],
                        best_votes.get(score["position"], ""),
                    ])
        return _workbook_response(workbook, "judge-records.xlsx")
